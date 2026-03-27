# coding=utf-8
"""
基于openpyxl重构Excel操作工具类，支持.xlsx格式（替代原xlrd/xlutils）
"""
import openpyxl
import time
import os


class ExcelUtil:
    def __init__(self, excel_path=None, index=None):
        """
        初始化Excel操作类（强制要求传入excel_path）
        :param excel_path: Excel文件路径（仅支持.xlsx格式，必填）
        :param index: 工作表索引（默认0）
        :raises ValueError: 未传入excel_path或路径格式不合法时抛出
        """
        # 1. 强制校验：未传入文件路径直接抛异常，提示清晰
        if excel_path is None:
            raise ValueError("【ExcelUtil初始化失败】必须传入excel_path参数！请指定需要操作的.xlsx文件路径")

        # 2. 路径格式校验：仅支持.xlsx
        self.excel_path = os.path.abspath(excel_path)  # 转为绝对路径，避免相对路径问题
        if not self.excel_path.endswith('.xlsx'):
            raise ValueError(f"【ExcelUtil初始化失败】仅支持.xlsx格式文件！当前传入的路径：{self.excel_path}")

        # 3. 处理工作表索引
        self.sheet_index = index if index is not None else 0

        # 4. 初始化工作簿和工作表
        self.workbook = None
        self.worksheet = None
        self._load_workbook()

    def _load_workbook(self):
        """内部方法：加载Excel工作簿和指定工作表"""
        try:
            # 检查文件是否存在，不存在则新建
            if not os.path.exists(self.excel_path):
                self.workbook = openpyxl.Workbook()
                self.workbook.save(self.excel_path)
                print(f"文件不存在，已新建：{self.excel_path}")

            # 加载工作簿（read_only=False 支持读写）
            self.workbook = openpyxl.load_workbook(self.excel_path)
            # 获取指定索引的工作表
            self.worksheet = self.workbook.worksheets[self.sheet_index]
        except Exception as e:
            raise Exception(f"加载Excel失败：{str(e)}")

    # 获取excel数据，按照每行一个list，添加到一个大的list里面
    def get_data(self):
        """
        读取所有行数据，返回二维列表
        :return: [[行1数据], [行2数据], ...]
        """
        result = []
        # openpyxl用iter_rows读取所有行，values_only=True只返回值（不返回单元格对象）
        for row in self.worksheet.iter_rows(values_only=True):
            result.append(list(row))
        return result if result else None

    # 获取excel行数
    def get_lines(self):
        """
        获取工作表的有效行数
        :return: 行数（int）或None
        """
        rows = self.worksheet.max_row
        return rows if rows >= 1 else None

    # 获取单元格的数据
    def get_col_value(self, row, col):
        """
        获取指定单元格数据（注意：openpyxl的行/列从1开始，与xlrd（从0开始）兼容）
        :param row: 行号（原xlrd的0对应openpyxl的1）
        :param col: 列号（原xlrd的0对应openpyxl的1）
        :return: 单元格值或None
        """
        # 兼容原代码的0起始索引，转换为openpyxl的1起始
        target_row = row + 1
        target_col = col + 1

        if self.get_lines() >= target_row:
            data = self.worksheet.cell(row=target_row, column=target_col).value
            return data
        return None

    # 写入数据（替代原xlutils.copy的写入逻辑）
    def write_value(self, row, value):
        """
        写入数据到指定行的第10列（原逻辑：row行，9列 → 转换为openpyxl的10列）
        :param row: 行号（原xlrd的0起始）
        :param value: 要写入的值
        """
        # 兼容原代码的0起始行号，转换为openpyxl的1起始
        target_row = row + 1
        target_col = 10  # 原代码的9列（0起始）→ 10列（1起始）

        # 写入数据
        self.worksheet.cell(row=target_row, column=target_col, value=value)

        # 保存文件（openpyxl必须手动save才生效）
        self.workbook.save(self.excel_path)
        self.workbook.close()  # 关闭工作簿，避免占用
        time.sleep(1)
        # 重新加载工作簿，确保后续操作数据最新
        self._load_workbook()


if __name__ == '__main__':
    # 注意：原代码传入的是.csv，openpyxl不支持，需改为.xlsx文件
    # 示例1：操作.xlsx文件
    ex = ExcelUtil('../config/RYAccountData.xlsx')  # 改为xlsx格式
    print(ex.get_col_value(1, 1))  # 读取第1行第1列数据（原0,0逻辑）

    # 示例2：测试写入
    ex.write_value(5, "测试写入")

    # 示例3：若需处理CSV，参考以下代码（补充CSV支持）

    """
    import csv
    def read_csv(csv_path, row, col):
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
            return rows[row][col] if len(rows) > row and len(rows[row]) > col else None
    # 调用CSV读取
    print(read_csv('../config/RYAccountData.csv', 0, 0))
    """